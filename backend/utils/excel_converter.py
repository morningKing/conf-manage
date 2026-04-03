"""
Excel 文件与 Luckysheet 格式转换工具
"""
import json
import os
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime


def excel_to_luckysheet(file_path):
    """
    将 Excel 文件转换为 Luckysheet 格式

    Args:
        file_path: Excel 文件路径

    Returns:
        list: Luckysheet sheet 数据列表
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")

    # 加载工作簿
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.xls':
        # 旧版 xls 格式，使用 xlrd 读取后转换
        import xlrd
        workbook = xlrd.open_workbook(file_path)
        return _xlrd_to_luckysheet(workbook, file_path)
    else:
        # xlsx 格式
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        return _openpyxl_to_luckysheet(workbook, file_path)


def _openpyxl_to_luckysheet(workbook, file_path):
    """使用 openpyxl 转换"""
    sheets = []

    for sheet_index, sheet_name in enumerate(workbook.sheetnames):
        sheet = workbook[sheet_name]

        # 构建单元格数据
        cell_data = []
        merge_cells = []

        # 获取合并单元格信息
        for merge_range in sheet.merged_cells.ranges:
            merge_cells.append({
                'r': merge_range.min_row - 1,  # Luckysheet 使用 0-based 索引
                'c': merge_range.min_col - 1,
                'rs': merge_range.max_row - merge_range.min_row + 1,
                'cs': merge_range.max_col - merge_range.min_col + 1
            })

        # 遍历所有有数据的行
        max_row = sheet.max_row
        max_col = sheet.max_column

        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    cell_data.append({
                        'r': row_idx - 1,  # 0-based
                        'c': col_idx - 1,
                        'v': _convert_cell_value(cell.value)
                    })

        # 构建列宽信息
        col_info = []
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            col_dim = sheet.column_dimensions.get(col_letter)
            if col_dim and col_dim.width:
                col_info.append({
                    'w': int(col_dim.width * 8)  # 转换为像素宽度
                })

        # 构建行高信息
        row_info = []
        for row_idx in range(1, max_row + 1):
            row_dim = sheet.row_dimensions.get(row_idx)
            if row_dim and row_dim.height:
                row_info.append({
                    'h': int(row_dim.height)
                })

        sheets.append({
            'name': sheet_name,
            'index': sheet_index,
            'order': sheet_index,
            'status': 1 if sheet_index == 0 else 0,
            'celldata': cell_data,
            'config': {
                'merge': merge_cells if merge_cells else None,
                'columnlen': col_info if col_info else None,
                'rowlen': row_info if row_info else None
            },
            'data': [],  # Luckysheet 会自动从 celldata 生成
            'row': max_row,
            'column': max_col
        })

    return sheets


def _xlrd_to_luckysheet(workbook, file_path):
    """使用 xlrd 转换旧版 xls 文件"""
    sheets = []

    for sheet_index in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_index)
        cell_data = []

        for row_idx in range(sheet.nrows):
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                if cell.ctype != 0:  # 0 = empty
                    cell_data.append({
                        'r': row_idx,
                        'c': col_idx,
                        'v': _convert_xlrd_value(cell)
                    })

        sheets.append({
            'name': sheet.name,
            'index': sheet_index,
            'order': sheet_index,
            'status': 1 if sheet_index == 0 else 0,
            'celldata': cell_data,
            'config': {},
            'data': [],
            'row': sheet.nrows,
            'column': sheet.ncols
        })

    return sheets


def _convert_cell_value(value):
    """转换单元格值为 Luckysheet 格式"""
    if value is None:
        return {'v': '', 'm': ''}

    if isinstance(value, (int, float)):
        # 数字
        return {
            'v': value,
            'm': str(value),
            'ct': {'fa': 'General', 't': 'n'}
        }
    elif isinstance(value, datetime):
        # 日期
        return {
            'v': value.strftime('%Y-%m-%d'),
            'm': value.strftime('%Y-%m-%d'),
            'ct': {'fa': 'yyyy-mm-dd', 't': 'd'}
        }
    elif isinstance(value, str):
        return {
            'v': value,
            'm': value,
            'ct': {'fa': 'General', 't': 'g'}
        }
    else:
        return {
            'v': str(value),
            'm': str(value),
            'ct': {'fa': 'General', 't': 'g'}
        }


def _convert_xlrd_value(cell):
    """转换 xlrd 单元格值"""
    import xlrd

    if cell.ctype == xlrd.XL_CELL_TEXT:
        return {
            'v': cell.value,
            'm': cell.value,
            'ct': {'fa': 'General', 't': 'g'}
        }
    elif cell.ctype == xlrd.XL_CELL_NUMBER:
        return {
            'v': cell.value,
            'm': str(cell.value),
            'ct': {'fa': 'General', 't': 'n'}
        }
    elif cell.ctype == xlrd.XL_CELL_DATE:
        date_tuple = xlrd.xldate_as_tuple(cell.value, 0)
        date_str = '%04d-%02d-%02d' % date_tuple[:3]
        return {
            'v': date_str,
            'm': date_str,
            'ct': {'fa': 'yyyy-mm-dd', 't': 'd'}
        }
    else:
        return {
            'v': str(cell.value),
            'm': str(cell.value)
        }


def luckysheet_to_excel(grid_data, output_path):
    """
    将 Luckysheet 数据保存为 Excel 文件

    Args:
        grid_data: Luckysheet sheet 数据列表
        output_path: 输出文件路径
    """
    workbook = openpyxl.Workbook()

    # 删除默认工作表
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    for sheet_data in grid_data:
        sheet_name = sheet_data.get('name', f'Sheet{sheet_data.get("index", 0) + 1}')
        sheet = workbook.create_sheet(title=sheet_name)

        # 处理单元格数据
        cell_data = sheet_data.get('celldata', [])
        for cell in cell_data:
            row = cell.get('r', 0) + 1  # openpyxl 使用 1-based 索引
            col = cell.get('c', 0) + 1
            value_info = cell.get('v', {})

            # 提取值
            if isinstance(value_info, dict):
                value = value_info.get('v', '')
            else:
                value = value_info

            sheet.cell(row=row, column=col, value=value)

        # 处理合并单元格
        config = sheet_data.get('config', {})
        merges = config.get('merge', [])
        if merges:
            for merge in merges:
                start_row = merge.get('r', 0) + 1
                start_col = merge.get('c', 0) + 1
                end_row = start_row + merge.get('rs', 1) - 1
                end_col = start_col + merge.get('cs', 1) - 1
                sheet.merge_cells(
                    start_row=start_row,
                    start_column=start_col,
                    end_row=end_row,
                    end_column=end_col
                )

        # 处理列宽
        col_len = config.get('columnlen', [])
        if col_len:
            for idx, col_info in enumerate(col_len):
                if col_info.get('w'):
                    col_letter = get_column_letter(idx + 1)
                    sheet.column_dimensions[col_letter].width = col_info['w'] / 8

    # 保存文件
    workbook.save(output_path)
    return output_path


def get_excel_info(file_path):
    """
    获取 Excel 文件基本信息

    Args:
        file_path: Excel 文件路径

    Returns:
        dict: 文件信息
    """
    if not os.path.exists(file_path):
        return None

    stat = os.stat(file_path)

    try:
        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.xls':
            import xlrd
            workbook = xlrd.open_workbook(file_path)
            return {
                'filename': os.path.basename(file_path),
                'size': stat.st_size,
                'sheets': workbook.sheet_names(),
                'sheet_count': workbook.nsheets
            }
        else:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            return {
                'filename': os.path.basename(file_path),
                'size': stat.st_size,
                'sheets': workbook.sheetnames,
                'sheet_count': len(workbook.sheetnames)
            }
    except Exception as e:
        return {
            'filename': os.path.basename(file_path),
            'size': stat.st_size,
            'error': str(e)
        }
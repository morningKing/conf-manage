"""
Excel 文件操作 API
提供 Excel 文件的读取、编辑、保存等功能
"""
from flask import request, jsonify
from . import api_bp
from config import Config
import os
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from utils import safe_filename

# 文件大小限制 (10MB)
MAX_EXCEL_SIZE = 10 * 1024 * 1024

# 分页限制
MAX_ROWS_PER_REQUEST = 100
MAX_COLS_PER_REQUEST = 50


def get_excel_path(file_path):
    """
    解析 Excel 文件路径
    支持上传目录和执行空间目录的文件

    Args:
        file_path: 相对路径或绝对路径

    Returns:
        tuple: (绝对路径, 错误信息) - 成功时错误信息为 None
    """
    if not file_path:
        return None, "文件路径不能为空"

    # 安全检查：防止路径遍历攻击
    if '..' in file_path or file_path.startswith('/') or ':' in file_path.split('/')[0] if '/' in file_path else ':' in file_path:
        # 允许 Windows 盘符路径，但需要进一步验证
        pass

    # 尝试不同的基础目录
    base_dirs = [
        Config.UPLOAD_FOLDER,  # 上传目录
        Config.EXECUTION_SPACES_DIR,  # 执行空间目录
        Config.WORKFLOW_EXECUTION_SPACES_DIR,  # 工作流执行空间目录
    ]

    for base_dir in base_dirs:
        full_path = os.path.join(base_dir, file_path)
        if os.path.exists(full_path):
            # 安全检查：确保路径在允许的目录内
            abs_path = os.path.abspath(full_path)
            abs_base = os.path.abspath(base_dir)
            if abs_path.startswith(abs_base):
                return abs_path, None

    # 如果提供了绝对路径，检查是否有效
    if os.path.isabs(file_path) and os.path.exists(file_path):
        # 验证是否在允许的目录内
        abs_path = os.path.abspath(file_path)
        for base_dir in base_dirs:
            abs_base = os.path.abspath(base_dir)
            if abs_path.startswith(abs_base):
                return abs_path, None
        return None, "文件路径不在允许的目录范围内"

    return None, "文件不存在"


def validate_excel_file(file_path):
    """
    验证 Excel 文件

    Args:
        file_path: 文件绝对路径

    Returns:
        tuple: (是否有效, 错误信息)
    """
    if not os.path.exists(file_path):
        return False, "文件不存在"

    if not os.path.isfile(file_path):
        return False, "不是有效的文件"

    # 检查文件大小
    file_size = os.path.getsize(file_path)
    if file_size > MAX_EXCEL_SIZE:
        return False, f"文件大小超过限制 ({file_size / 1024 / 1024:.2f}MB > {MAX_EXCEL_SIZE / 1024 / 1024:.0f}MB)"

    # 检查文件扩展名
    ext = os.path.splitext(file_path)[1].lower()
    if ext not in ['.xlsx', '.xls']:
        return False, "不支持的文件格式，仅支持 .xlsx 和 .xls 文件"

    return True, None


@api_bp.route('/excel/info', methods=['GET'])
def get_excel_file_info():
    """
    获取 Excel 文件信息
    返回文件的基本信息：工作表列表、行数、列数、文件大小等

    Query Parameters:
        path: 文件路径（相对路径）

    Returns:
        {
            "code": 0,
            "data": {
                "filename": "文件名.xlsx",
                "size": 文件大小(字节),
                "sheets": [
                    {
                        "name": "Sheet1",
                        "index": 0,
                        "rows": 100,
                        "cols": 20
                    }
                ],
                "sheet_count": 1
            }
        }
    """
    try:
        file_path = request.args.get('path', '')

        # 解析文件路径
        full_path, error = get_excel_path(file_path)
        if error:
            return jsonify({'code': 1, 'message': error}), 400

        # 验证文件
        is_valid, error = validate_excel_file(full_path)
        if not is_valid:
            return jsonify({'code': 1, 'message': error}), 400

        # 获取文件信息
        stat = os.stat(full_path)
        ext = os.path.splitext(full_path)[1].lower()

        sheets_info = []

        if ext == '.xls':
            # 旧版 xls 格式
            try:
                import xlrd
                workbook = xlrd.open_workbook(full_path)
                for idx in range(workbook.nsheets):
                    sheet = workbook.sheet_by_index(idx)
                    sheets_info.append({
                        'name': sheet.name,
                        'index': idx,
                        'rows': sheet.nrows,
                        'cols': sheet.ncols
                    })
                sheet_count = workbook.nsheets
            except ImportError:
                return jsonify({'code': 1, 'message': '服务器未安装 xlrd 库，不支持 .xls 格式'}), 500
        else:
            # xlsx 格式
            workbook = openpyxl.load_workbook(full_path, read_only=True)
            for idx, sheet_name in enumerate(workbook.sheetnames):
                sheet = workbook[sheet_name]
                sheets_info.append({
                    'name': sheet_name,
                    'index': idx,
                    'rows': sheet.max_row,
                    'cols': sheet.max_column
                })
            sheet_count = len(workbook.sheetnames)
            workbook.close()

        return jsonify({
            'code': 0,
            'data': {
                'filename': os.path.basename(full_path),
                'size': stat.st_size,
                'sheets': sheets_info,
                'sheet_count': sheet_count,
                'path': file_path
            }
        })

    except Exception as e:
        return jsonify({'code': 1, 'message': f'获取文件信息失败: {str(e)}'}), 500


@api_bp.route('/excel/sheet', methods=['GET'])
def get_excel_sheet_data():
    """
    获取 Excel 工作表数据（分页）

    Query Parameters:
        path: 文件路径
        sheet: 工作表名称或索引（默认第一个工作表）
        offset: 起始行号（从0开始，默认0）
        limit: 返回行数（默认100，最大100）
        cols: 返回列数（默认50，最大50）

    Returns:
        {
            "code": 0,
            "data": {
                "sheet_name": "Sheet1",
                "total_rows": 1000,
                "total_cols": 20,
                "offset": 0,
                "limit": 100,
                "rows": [[...], ...],
                "merge_cells": [{"r": 0, "c": 0, "rs": 2, "cs": 2}, ...]
            }
        }
    """
    try:
        file_path = request.args.get('path', '')
        sheet_identifier = request.args.get('sheet', '0')
        offset = request.args.get('offset', 0, type=int)
        limit = min(request.args.get('limit', MAX_ROWS_PER_REQUEST, type=int), MAX_ROWS_PER_REQUEST)
        max_cols = min(request.args.get('cols', MAX_COLS_PER_REQUEST, type=int), MAX_COLS_PER_REQUEST)

        # 解析文件路径
        full_path, error = get_excel_path(file_path)
        if error:
            return jsonify({'code': 1, 'message': error}), 400

        # 验证文件
        is_valid, error = validate_excel_file(full_path)
        if not is_valid:
            return jsonify({'code': 1, 'message': error}), 400

        # 确保参数有效
        if offset < 0:
            offset = 0
        if limit < 1:
            limit = MAX_ROWS_PER_REQUEST

        ext = os.path.splitext(full_path)[1].lower()

        if ext == '.xls':
            # 旧版 xls 格式
            try:
                import xlrd
                workbook = xlrd.open_workbook(full_path)

                # 获取工作表
                if sheet_identifier.isdigit():
                    sheet = workbook.sheet_by_index(int(sheet_identifier))
                else:
                    sheet = workbook.sheet_by_name(sheet_identifier)

                total_rows = sheet.nrows
                total_cols = min(sheet.ncols, max_cols)

                # 读取数据
                rows = []
                end_row = min(offset + limit, total_rows)
                for row_idx in range(offset, end_row):
                    row_data = []
                    for col_idx in range(total_cols):
                        cell = sheet.cell(row_idx, col_idx)
                        if cell.ctype == 0:  # empty
                            row_data.append(None)
                        elif cell.ctype == 2:  # number
                            row_data.append(cell.value)
                        elif cell.ctype == 3:  # date
                            date_tuple = xlrd.xldate_as_tuple(cell.value, 0)
                            row_data.append('%04d-%02d-%02d' % date_tuple[:3])
                        else:
                            row_data.append(cell.value)
                    rows.append(row_data)

                # xls 格式不支持合并单元格信息，返回空列表
                merge_cells = []

                return jsonify({
                    'code': 0,
                    'data': {
                        'sheet_name': sheet.name,
                        'total_rows': total_rows,
                        'total_cols': sheet.ncols,
                        'offset': offset,
                        'limit': limit,
                        'rows': rows,
                        'merge_cells': merge_cells
                    }
                })

            except ImportError:
                return jsonify({'code': 1, 'message': '服务器未安装 xlrd 库，不支持 .xls 格式'}), 500
        else:
            # xlsx 格式
            workbook = openpyxl.load_workbook(full_path, read_only=True, data_only=True)

            # 获取工作表
            if sheet_identifier.isdigit():
                sheet_idx = int(sheet_identifier)
                if sheet_idx < 0 or sheet_idx >= len(workbook.sheetnames):
                    workbook.close()
                    return jsonify({'code': 1, 'message': '工作表索引超出范围'}), 400
                sheet_name = workbook.sheetnames[sheet_idx]
                sheet = workbook[sheet_name]
            else:
                if sheet_identifier not in workbook.sheetnames:
                    workbook.close()
                    return jsonify({'code': 1, 'message': f'工作表 "{sheet_identifier}" 不存在'}), 400
                sheet = workbook[sheet_identifier]
                sheet_name = sheet_identifier

            total_rows = sheet.max_row
            total_cols = sheet.max_column

            # 读取数据
            rows = []
            end_row = min(offset + limit + 1, total_rows + 1)  # openpyxl 使用 1-based 索引
            actual_cols = min(total_cols, max_cols)

            for row_idx in range(offset + 1, end_row):  # openpyxl 使用 1-based 索引
                row_data = []
                for col_idx in range(1, actual_cols + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    # 处理日期类型
                    if isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    row_data.append(value)
                rows.append(row_data)

            # 获取合并单元格信息
            merge_cells = []
            for merge_range in sheet.merged_cells.ranges:
                merge_cells.append({
                    'r': merge_range.min_row - 1,  # 转换为 0-based 索引
                    'c': merge_range.min_col - 1,
                    'rs': merge_range.max_row - merge_range.min_row + 1,
                    'cs': merge_range.max_col - merge_range.min_col + 1
                })

            workbook.close()

            return jsonify({
                'code': 0,
                'data': {
                    'sheet_name': sheet_name,
                    'total_rows': total_rows,
                    'total_cols': total_cols,
                    'offset': offset,
                    'limit': limit,
                    'rows': rows,
                    'merge_cells': merge_cells
                }
            })

    except Exception as e:
        return jsonify({'code': 1, 'message': f'读取工作表数据失败: {str(e)}'}), 500


@api_bp.route('/excel/save', methods=['POST'])
def save_excel_editor_file():
    """
    保存 Excel 文件
    支持增量更新和完整保存

    Request Body:
        {
            "path": "文件路径",
            "sheet": "工作表名称或索引",
            "data": {
                "cells": [  // 增量更新：只更新指定的单元格
                    {"row": 0, "col": 0, "value": "新值"},
                    ...
                ],
                // 或者
                "rows": [[...], ...]  // 完整更新：替换整个工作表数据
            },
            "merge_cells": [  // 可选：合并单元格信息
                {"r": 0, "c": 0, "rs": 2, "cs": 2},
                ...
            ]
        }

    Returns:
        {
            "code": 0,
            "message": "保存成功"
        }
    """
    try:
        data = request.get_json()
        file_path = data.get('path', '')
        sheet_identifier = data.get('sheet', '0')
        cells = data.get('data', {}).get('cells', [])
        rows = data.get('data', {}).get('rows', None)
        merge_cells = data.get('merge_cells', None)

        if not file_path:
            return jsonify({'code': 1, 'message': '文件路径不能为空'}), 400

        # 解析文件路径
        full_path, error = get_excel_path(file_path)
        if error:
            return jsonify({'code': 1, 'message': error}), 400

        # 验证文件
        is_valid, error = validate_excel_file(full_path)
        if not is_valid:
            return jsonify({'code': 1, 'message': error}), 400

        ext = os.path.splitext(full_path)[1].lower()

        if ext == '.xls':
            return jsonify({'code': 1, 'message': '.xls 格式文件暂不支持编辑，请转换为 .xlsx 格式'}), 400

        # 加载工作簿
        workbook = openpyxl.load_workbook(full_path)

        # 获取工作表
        if sheet_identifier.isdigit():
            sheet_idx = int(sheet_identifier)
            if sheet_idx < 0 or sheet_idx >= len(workbook.sheetnames):
                workbook.close()
                return jsonify({'code': 1, 'message': '工作表索引超出范围'}), 400
            sheet = workbook.worksheets[sheet_idx]
            sheet_name = sheet.title
        else:
            if sheet_identifier not in workbook.sheetnames:
                workbook.close()
                return jsonify({'code': 1, 'message': f'工作表 "{sheet_identifier}" 不存在'}), 400
            sheet = workbook[sheet_identifier]
            sheet_name = sheet_identifier

        # 增量更新：更新指定单元格
        if cells:
            for cell_data in cells:
                row = cell_data.get('row', 0)
                col = cell_data.get('col', 0)
                value = cell_data.get('value')
                # openpyxl 使用 1-based 索引
                sheet.cell(row=row + 1, column=col + 1, value=value)

        # 完整更新：替换整个工作表数据
        elif rows is not None:
            # 清除现有数据
            for row in sheet.iter_rows():
                for cell in row:
                    cell.value = None

            # 写入新数据
            for row_idx, row_data in enumerate(rows):
                for col_idx, value in enumerate(row_data):
                    sheet.cell(row=row_idx + 1, column=col_idx + 1, value=value)

        # 更新合并单元格
        if merge_cells is not None:
            # 先清除现有合并单元格
            existing_merges = list(sheet.merged_cells.ranges)
            for merge_range in existing_merges:
                sheet.unmerge_cells(str(merge_range))

            # 添加新的合并单元格
            for merge in merge_cells:
                start_row = merge.get('r', 0) + 1
                start_col = merge.get('c', 0) + 1
                end_row = start_row + merge.get('rs', 1) - 1
                end_col = start_col + merge.get('cs', 1) - 1
                sheet.merge_cells(start_row=start_row, start_column=start_col,
                                  end_row=end_row, end_column=end_col)

        # 保存文件
        workbook.save(full_path)
        workbook.close()

        return jsonify({
            'code': 0,
            'message': '保存成功'
        })

    except Exception as e:
        return jsonify({'code': 1, 'message': f'保存文件失败: {str(e)}'}), 500


@api_bp.route('/excel/sheet/add', methods=['POST'])
def add_excel_sheet():
    """
    添加新工作表

    Request Body:
        {
            "path": "文件路径",
            "name": "新工作表名称",
            "position": 0  // 可选：插入位置（默认在最后）
        }

    Returns:
        {
            "code": 0,
            "data": {
                "name": "新工作表名称",
                "index": 1
            },
            "message": "工作表添加成功"
        }
    """
    try:
        data = request.get_json()
        file_path = data.get('path', '')
        sheet_name = data.get('name', '')
        position = data.get('position')

        if not file_path:
            return jsonify({'code': 1, 'message': '文件路径不能为空'}), 400

        if not sheet_name:
            return jsonify({'code': 1, 'message': '工作表名称不能为空'}), 400

        # 处理工作表名称
        sheet_name = safe_filename(sheet_name)
        # 移除扩展名（如果有）
        if '.' in sheet_name:
            sheet_name = sheet_name.split('.')[0]

        # 解析文件路径
        full_path, error = get_excel_path(file_path)
        if error:
            return jsonify({'code': 1, 'message': error}), 400

        # 验证文件
        is_valid, error = validate_excel_file(full_path)
        if not is_valid:
            return jsonify({'code': 1, 'message': error}), 400

        ext = os.path.splitext(full_path)[1].lower()

        if ext == '.xls':
            return jsonify({'code': 1, 'message': '.xls 格式文件暂不支持编辑，请转换为 .xlsx 格式'}), 400

        # 加载工作簿
        workbook = openpyxl.load_workbook(full_path)

        # 检查工作表名称是否已存在
        if sheet_name in workbook.sheetnames:
            workbook.close()
            return jsonify({'code': 1, 'message': f'工作表 "{sheet_name}" 已存在'}), 400

        # 创建新工作表
        if position is not None and isinstance(position, int) and 0 <= position <= len(workbook.sheetnames):
            new_sheet = workbook.create_sheet(title=sheet_name, index=position)
            sheet_index = position
        else:
            new_sheet = workbook.create_sheet(title=sheet_name)
            sheet_index = len(workbook.sheetnames) - 1

        # 保存文件
        workbook.save(full_path)
        workbook.close()

        return jsonify({
            'code': 0,
            'data': {
                'name': sheet_name,
                'index': sheet_index
            },
            'message': '工作表添加成功'
        })

    except Exception as e:
        return jsonify({'code': 1, 'message': f'添加工作表失败: {str(e)}'}), 500


@api_bp.route('/excel/sheet/delete', methods=['DELETE'])
def delete_excel_sheet():
    """
    删除工作表

    Query Parameters:
        path: 文件路径
        sheet: 工作表名称或索引

    Returns:
        {
            "code": 0,
            "message": "工作表删除成功"
        }
    """
    try:
        file_path = request.args.get('path', '')
        sheet_identifier = request.args.get('sheet', '')

        if not file_path:
            return jsonify({'code': 1, 'message': '文件路径不能为空'}), 400

        if not sheet_identifier:
            return jsonify({'code': 1, 'message': '工作表名称或索引不能为空'}), 400

        # 解析文件路径
        full_path, error = get_excel_path(file_path)
        if error:
            return jsonify({'code': 1, 'message': error}), 400

        # 验证文件
        is_valid, error = validate_excel_file(full_path)
        if not is_valid:
            return jsonify({'code': 1, 'message': error}), 400

        ext = os.path.splitext(full_path)[1].lower()

        if ext == '.xls':
            return jsonify({'code': 1, 'message': '.xls 格式文件暂不支持编辑，请转换为 .xlsx 格式'}), 400

        # 加载工作簿
        workbook = openpyxl.load_workbook(full_path)

        # 检查是否只有一个工作表
        if len(workbook.sheetnames) <= 1:
            workbook.close()
            return jsonify({'code': 1, 'message': '无法删除最后一个工作表'}), 400

        # 获取工作表
        if sheet_identifier.isdigit():
            sheet_idx = int(sheet_identifier)
            if sheet_idx < 0 or sheet_idx >= len(workbook.sheetnames):
                workbook.close()
                return jsonify({'code': 1, 'message': '工作表索引超出范围'}), 400
            sheet_name = workbook.sheetnames[sheet_idx]
        else:
            if sheet_identifier not in workbook.sheetnames:
                workbook.close()
                return jsonify({'code': 1, 'message': f'工作表 "{sheet_identifier}" 不存在'}), 400
            sheet_name = sheet_identifier

        # 删除工作表
        del workbook[sheet_name]

        # 保存文件
        workbook.save(full_path)
        workbook.close()

        return jsonify({
            'code': 0,
            'message': '工作表删除成功'
        })

    except Exception as e:
        return jsonify({'code': 1, 'message': f'删除工作表失败: {str(e)}'}), 500


@api_bp.route('/excel/sheet/rename', methods=['PUT'])
def rename_excel_sheet():
    """
    重命名工作表

    Request Body:
        {
            "path": "文件路径",
            "sheet": "工作表名称或索引",
            "new_name": "新工作表名称"
        }

    Returns:
        {
            "code": 0,
            "message": "工作表重命名成功"
        }
    """
    try:
        data = request.get_json()
        file_path = data.get('path', '')
        sheet_identifier = data.get('sheet', '')
        new_name = data.get('new_name', '')

        if not file_path:
            return jsonify({'code': 1, 'message': '文件路径不能为空'}), 400

        if not sheet_identifier:
            return jsonify({'code': 1, 'message': '工作表名称或索引不能为空'}), 400

        if not new_name:
            return jsonify({'code': 1, 'message': '新工作表名称不能为空'}), 400

        # 处理工作表名称
        new_name = safe_filename(new_name)
        # 移除扩展名（如果有）
        if '.' in new_name:
            new_name = new_name.split('.')[0]

        # 解析文件路径
        full_path, error = get_excel_path(file_path)
        if error:
            return jsonify({'code': 1, 'message': error}), 400

        # 验证文件
        is_valid, error = validate_excel_file(full_path)
        if not is_valid:
            return jsonify({'code': 1, 'message': error}), 400

        ext = os.path.splitext(full_path)[1].lower()

        if ext == '.xls':
            return jsonify({'code': 1, 'message': '.xls 格式文件暂不支持编辑，请转换为 .xlsx 格式'}), 400

        # 加载工作簿
        workbook = openpyxl.load_workbook(full_path)

        # 检查新名称是否已存在
        if new_name in workbook.sheetnames:
            workbook.close()
            return jsonify({'code': 1, 'message': f'工作表 "{new_name}" 已存在'}), 400

        # 获取工作表
        if sheet_identifier.isdigit():
            sheet_idx = int(sheet_identifier)
            if sheet_idx < 0 or sheet_idx >= len(workbook.sheetnames):
                workbook.close()
                return jsonify({'code': 1, 'message': '工作表索引超出范围'}), 400
            sheet = workbook.worksheets[sheet_idx]
        else:
            if sheet_identifier not in workbook.sheetnames:
                workbook.close()
                return jsonify({'code': 1, 'message': f'工作表 "{sheet_identifier}" 不存在'}), 400
            sheet = workbook[sheet_identifier]

        # 重命名工作表
        sheet.title = new_name

        # 保存文件
        workbook.save(full_path)
        workbook.close()

        return jsonify({
            'code': 0,
            'message': '工作表重命名成功'
        })

    except Exception as e:
        return jsonify({'code': 1, 'message': f'重命名工作表失败: {str(e)}'}), 500
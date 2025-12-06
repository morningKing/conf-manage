"""
文件管理API
"""
from flask import request, jsonify, send_file, send_from_directory
from werkzeug.utils import secure_filename
from . import api_bp
import os
from config import Config


@api_bp.route('/files', methods=['GET'])
def list_files():
    """获取文件列表"""
    try:
        path = request.args.get('path', '')
        base_dir = os.path.join(Config.DATA_DIR, 'uploads')
        target_dir = os.path.join(base_dir, path)

        # 安全检查
        if not os.path.abspath(target_dir).startswith(os.path.abspath(base_dir)):
            return jsonify({'code': 1, 'message': '非法路径'}), 400

        if not os.path.exists(target_dir):
            os.makedirs(target_dir, exist_ok=True)

        files = []
        for item in os.listdir(target_dir):
            item_path = os.path.join(target_dir, item)
            relative_path = os.path.join(path, item)

            files.append({
                'name': item,
                'path': relative_path,
                'is_dir': os.path.isdir(item_path),
                'size': os.path.getsize(item_path) if os.path.isfile(item_path) else 0,
                'modified_at': os.path.getmtime(item_path)
            })

        return jsonify({
            'code': 0,
            'data': sorted(files, key=lambda x: (not x['is_dir'], x['name']))
        })
    except Exception as e:
        return jsonify({'code': 1, 'message': str(e)}), 500


@api_bp.route('/files/upload', methods=['POST'])
def upload_file():
    """上传文件"""
    try:
        if 'file' not in request.files:
            return jsonify({'code': 1, 'message': '没有文件'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'code': 1, 'message': '文件名为空'}), 400

        path = request.form.get('path', '')
        filename = secure_filename(file.filename)

        base_dir = os.path.join(Config.DATA_DIR, 'uploads')
        target_dir = os.path.join(base_dir, path)

        # 安全检查
        if not os.path.abspath(target_dir).startswith(os.path.abspath(base_dir)):
            return jsonify({'code': 1, 'message': '非法路径'}), 400

        os.makedirs(target_dir, exist_ok=True)

        file_path = os.path.join(target_dir, filename)
        file.save(file_path)

        return jsonify({
            'code': 0,
            'data': {
                'filename': filename,
                'path': os.path.join(path, filename)
            },
            'message': '文件上传成功'
        })
    except Exception as e:
        return jsonify({'code': 1, 'message': str(e)}), 500


@api_bp.route('/files/download', methods=['GET'])
def download_file():
    """下载文件"""
    try:
        path = request.args.get('path', '')
        base_dir = os.path.join(Config.DATA_DIR, 'uploads')
        file_path = os.path.join(base_dir, path)

        # 安全检查
        if not os.path.abspath(file_path).startswith(os.path.abspath(base_dir)):
            return jsonify({'code': 1, 'message': '非法路径'}), 400

        if not os.path.exists(file_path) or not os.path.isfile(file_path):
            return jsonify({'code': 1, 'message': '文件不存在'}), 404

        return send_file(file_path, as_attachment=True)
    except Exception as e:
        return jsonify({'code': 1, 'message': str(e)}), 500


@api_bp.route('/files/preview', methods=['GET'])
def preview_file():
    """预览文件"""
    try:
        path = request.args.get('path', '')
        base_dir = os.path.join(Config.DATA_DIR, 'uploads')
        file_path = os.path.join(base_dir, path)

        # 安全检查
        if not os.path.abspath(file_path).startswith(os.path.abspath(base_dir)):
            return jsonify({'code': 1, 'message': '非法路径'}), 400

        if not os.path.exists(file_path) or not os.path.isfile(file_path):
            return jsonify({'code': 1, 'message': '文件不存在'}), 404

        # 获取文件扩展名
        ext = os.path.splitext(file_path)[1].lower()

        # 文本文件预览 (txt, md, log, py, js, etc.)
        if ext in ['.txt', '.md', '.log', '.py', '.js', '.json', '.xml', '.html', '.css', '.yaml', '.yml', '.ini', '.conf', '.sh', '.bat', '.csv']:
            try:
                # 尝试不同的编码
                encodings = ['utf-8', 'gbk', 'gb2312', 'latin1']
                content = None
                used_encoding = None

                for encoding in encodings:
                    try:
                        with open(file_path, 'r', encoding=encoding) as f:
                            content = f.read(1024 * 1024)  # 最多读取1MB
                        used_encoding = encoding
                        break
                    except UnicodeDecodeError:
                        continue

                if content is None:
                    return jsonify({'code': 1, 'message': '无法读取文件内容，可能是二进制文件或编码不支持'}), 400

                return jsonify({
                    'code': 0,
                    'data': {
                        'type': 'text',
                        'content': content,
                        'encoding': used_encoding,
                        'extension': ext
                    }
                })
            except Exception as e:
                return jsonify({'code': 1, 'message': f'文本文件读取失败: {str(e)}'}), 500

        # 图片文件预览 (jpg, png, gif, bmp, webp, svg)
        elif ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.svg', '.ico']:
            try:
                import base64
                with open(file_path, 'rb') as f:
                    image_data = f.read()
                base64_data = base64.b64encode(image_data).decode('utf-8')

                # 确定MIME类型
                mime_types = {
                    '.jpg': 'image/jpeg',
                    '.jpeg': 'image/jpeg',
                    '.png': 'image/png',
                    '.gif': 'image/gif',
                    '.bmp': 'image/bmp',
                    '.webp': 'image/webp',
                    '.svg': 'image/svg+xml',
                    '.ico': 'image/x-icon'
                }
                mime_type = mime_types.get(ext, 'image/jpeg')

                return jsonify({
                    'code': 0,
                    'data': {
                        'type': 'image',
                        'content': base64_data,
                        'mime_type': mime_type,
                        'extension': ext
                    }
                })
            except Exception as e:
                return jsonify({'code': 1, 'message': f'图片文件读取失败: {str(e)}'}), 500

        # JSON文件特殊处理（格式化显示）
        elif ext == '.json':
            try:
                import json
                with open(file_path, 'r', encoding='utf-8') as f:
                    json_data = json.load(f)

                return jsonify({
                    'code': 0,
                    'data': {
                        'type': 'json',
                        'content': json.dumps(json_data, ensure_ascii=False, indent=2),
                        'json_data': json_data
                    }
                })
            except Exception as e:
                # 如果JSON解析失败，尝试作为普通文本读取
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read(1024 * 1024)
                    return jsonify({
                        'code': 0,
                        'data': {
                            'type': 'text',
                            'content': content,
                            'encoding': 'utf-8',
                            'extension': ext
                        }
                    })
                except:
                    return jsonify({'code': 1, 'message': f'JSON文件读取失败: {str(e)}'}), 500

        # Excel文件预览
        elif ext in ['.xlsx', '.xls']:
            import openpyxl
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheets_data = []

                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    rows = []
                    max_rows = min(sheet.max_row, 100)  # 限制最多100行
                    max_cols = min(sheet.max_column, 50)  # 限制最多50列

                    for row in sheet.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
                        # 将None转换为空字符串,其他值转为字符串
                        row_data = [str(cell) if cell is not None else '' for cell in row]
                        rows.append(row_data)

                    sheets_data.append({
                        'name': sheet_name,
                        'rows': rows
                    })

                wb.close()

                return jsonify({
                    'code': 0,
                    'data': {
                        'type': 'excel',
                        'sheets': sheets_data
                    }
                })
            except Exception as e:
                return jsonify({'code': 1, 'message': f'Excel文件解析失败: {str(e)}'}), 500

        # PDF文件预览
        elif ext == '.pdf':
            try:
                import base64
                with open(file_path, 'rb') as f:
                    pdf_data = f.read()
                base64_data = base64.b64encode(pdf_data).decode('utf-8')

                return jsonify({
                    'code': 0,
                    'data': {
                        'type': 'pdf',
                        'content': base64_data
                    }
                })
            except Exception as e:
                return jsonify({'code': 1, 'message': f'PDF文件读取失败: {str(e)}'}), 500

        else:
            return jsonify({'code': 1, 'message': f'不支持的文件类型: {ext}'}), 400

    except Exception as e:
        return jsonify({'code': 1, 'message': str(e)}), 500


@api_bp.route('/files/delete', methods=['DELETE'])
def delete_file():
    """删除文件"""
    try:
        path = request.args.get('path', '')
        base_dir = os.path.join(Config.DATA_DIR, 'uploads')
        file_path = os.path.join(base_dir, path)

        # 安全检查
        if not os.path.abspath(file_path).startswith(os.path.abspath(base_dir)):
            return jsonify({'code': 1, 'message': '非法路径'}), 400

        if not os.path.exists(file_path):
            return jsonify({'code': 1, 'message': '文件不存在'}), 404

        if os.path.isfile(file_path):
            os.remove(file_path)
        elif os.path.isdir(file_path):
            import shutil
            shutil.rmtree(file_path)

        return jsonify({
            'code': 0,
            'message': '删除成功'
        })
    except Exception as e:
        return jsonify({'code': 1, 'message': str(e)}), 500


@api_bp.route('/files/create-folder', methods=['POST'])
def create_folder():
    """创建文件夹"""
    try:
        data = request.get_json()
        path = data.get('path', '')
        folder_name = data.get('name', '')

        if not folder_name:
            return jsonify({'code': 1, 'message': '文件夹名称不能为空'}), 400

        folder_name = secure_filename(folder_name)
        base_dir = os.path.join(Config.DATA_DIR, 'uploads')
        target_path = os.path.join(base_dir, path, folder_name)

        # 安全检查
        if not os.path.abspath(target_path).startswith(os.path.abspath(base_dir)):
            return jsonify({'code': 1, 'message': '非法路径'}), 400

        if os.path.exists(target_path):
            return jsonify({'code': 1, 'message': '文件夹已存在'}), 400

        os.makedirs(target_path)

        return jsonify({
            'code': 0,
            'data': {
                'name': folder_name,
                'path': os.path.join(path, folder_name)
            },
            'message': '文件夹创建成功'
        })
    except Exception as e:
        return jsonify({'code': 1, 'message': str(e)}), 500


@api_bp.route('/files/update', methods=['PUT'])
def update_file():
    """更新文件内容"""
    try:
        data = request.get_json()
        path = data.get('path', '')
        content = data.get('content', '')

        if not path:
            return jsonify({'code': 1, 'message': '文件路径不能为空'}), 400

        base_dir = os.path.join(Config.DATA_DIR, 'uploads')
        file_path = os.path.join(base_dir, path)

        # 安全检查
        if not os.path.abspath(file_path).startswith(os.path.abspath(base_dir)):
            return jsonify({'code': 1, 'message': '非法路径'}), 400

        if not os.path.exists(file_path):
            return jsonify({'code': 1, 'message': '文件不存在'}), 404

        if not os.path.isfile(file_path):
            return jsonify({'code': 1, 'message': '不是文件'}), 400

        # 检查文件类型是否可编辑
        ext = os.path.splitext(file_path)[1].lower()
        editable_extensions = [
            '.txt', '.md', '.log', '.py', '.js', '.json', '.xml', '.html',
            '.css', '.yaml', '.yml', '.ini', '.conf', '.sh', '.bat', '.csv', '.sql'
        ]

        if ext not in editable_extensions:
            return jsonify({'code': 1, 'message': f'不支持编辑该类型的文件: {ext}'}), 400

        # 备份原文件
        backup_path = file_path + '.backup'
        if os.path.exists(file_path):
            import shutil
            shutil.copy2(file_path, backup_path)

        try:
            # 写入新内容
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(content)

            # 删除备份
            if os.path.exists(backup_path):
                os.remove(backup_path)

            return jsonify({
                'code': 0,
                'message': '文件保存成功'
            })
        except Exception as e:
            # 恢复备份
            if os.path.exists(backup_path):
                import shutil
                shutil.move(backup_path, file_path)
            raise e

    except Exception as e:
        return jsonify({'code': 1, 'message': str(e)}), 500
